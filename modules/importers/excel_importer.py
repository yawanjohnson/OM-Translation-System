"""
excel_importer.py - 從 Excel/CSV 批次匯入多語言翻譯到資料庫

支援格式：
  A欄=ENG原文, B欄=GER, C欄=DUT, ... (橫向展開)
  第一列可以是語言代碼（如 ENG, GER ...）或中文標題（如 英文, 德文 ...）
  也支援額外欄位：product（產品型號）, chapter（章節）
"""
import csv
import os
import openpyxl

LANG_CODES = [
    'ENG', 'GER', 'DUT', 'DAN', 'FRE', 'SPA', 'ITA', 'GRK',
    'POL', 'PRB', 'RUS', 'CHT', 'JPN', 'KOR', 'VTM', 'THI',
    'ARB', 'TRK', 'CHS'
]

# 中文欄位標題對應（容錯）
LANG_ALIASES = {
    '英文': 'ENG', 'english': 'ENG', 'eng': 'ENG',
    '德文': 'GER', 'german': 'GER', 'deutsch': 'GER', 'ger': 'GER',
    '荷文': 'DUT', 'dutch': 'DUT', 'dut': 'DUT',
    '丹麥文': 'DAN', 'danish': 'DAN', 'dan': 'DAN',
    '法文': 'FRE', 'french': 'FRE', 'français': 'FRE', 'fre': 'FRE',
    '西班牙文': 'SPA', 'spanish': 'SPA', 'español': 'SPA', 'spa': 'SPA',
    '義大利文': 'ITA', 'italian': 'ITA', 'italiano': 'ITA', 'ita': 'ITA',
    '希臘文': 'GRK', 'greek': 'GRK', 'grk': 'GRK',
    '波蘭文': 'POL', 'polish': 'POL', 'pol': 'POL',
    '葡萄牙文': 'PRB', 'portuguese': 'PRB', 'português': 'PRB', 'prb': 'PRB',
    '俄文': 'RUS', 'russian': 'RUS', 'rus': 'RUS',
    '繁體中文': 'CHT', 'cht': 'CHT', '中文': 'CHT',
    '日文': 'JPN', 'japanese': 'JPN', 'jpn': 'JPN',
    '韓文': 'KOR', 'korean': 'KOR', 'kor': 'KOR',
    '越南文': 'VTM', 'vietnamese': 'VTM', 'vtm': 'VTM',
    '泰文': 'THI', 'thai': 'THI', 'thi': 'THI',
    '阿拉伯文': 'ARB', 'arabic': 'ARB', 'arb': 'ARB',
    '土耳其文': 'TRK', 'turkish': 'TRK', 'trk': 'TRK',
    '簡體中文': 'CHS', 'chs': 'CHS', '简体中文': 'CHS',
    '產品': 'product', 'product': 'product', '型號': 'product',
    '章節': 'chapter', 'chapter': 'chapter', '段落': 'chapter',
}


def import_excel(file_path: str) -> list[dict]:
    """
    讀取 Excel 或 CSV，回傳 list of dicts（每個 dict 對應一列資料）。
    dict 的 key 為標準語言代碼（ENG, GER...）或 product/chapter。
    """
    ext = os.path.splitext(file_path)[1].lower()
    if ext in ('.xlsx', '.xls'):
        return _read_xlsx(file_path)
    elif ext == '.csv':
        return _read_csv(file_path)
    else:
        raise ValueError(f'不支援的檔案格式：{ext}')


def _read_xlsx(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header_row = [str(c).strip() if c is not None else '' for c in rows[0]]
    col_map = _build_col_map(header_row)
    return _parse_rows(rows[1:], col_map)


def _read_csv(path: str) -> list[dict]:
    with open(path, encoding='utf-8-sig', newline='') as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return []
    header_row = [c.strip() for c in rows[0]]
    col_map = _build_col_map(header_row)
    return _parse_rows(rows[1:], col_map)


def _build_col_map(headers: list) -> dict:
    """
    將欄位標題映射到標準語言代碼或保留原始名稱。
    回傳 {col_index: mapped_key_or_original_key}
    """
    col_map = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        h_str = str(h).strip()
        key = h_str.lower()
        if not key:
            continue
        if key in LANG_ALIASES:
            col_map[i] = LANG_ALIASES[key]
        elif key.upper() in LANG_CODES:
            col_map[i] = key.upper()
        else:
            col_map[i] = h_str  # 保留原始欄位名稱，方便自訂表格欄位讀取
    return col_map


def _parse_rows(rows, col_map: dict) -> list[dict]:
    result = []
    for i, row in enumerate(rows, start=2):
        if not row:
            continue
        entry = {}
        entry['_row_num'] = i
        for col_idx, lang_code in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None and str(val).strip():
                entry[lang_code] = str(val).strip()
        
        # 至少有除了 _row_num 以外的其他有效資料
        has_data = any(k != '_row_num' for k in entry.keys())
        if has_data:
            result.append(entry)
    return result


def generate_template(output_path: str):
    """產生空白 Excel 範本供 PM 填寫修改指示。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '翻譯對照表'

    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill('solid', fgColor='1B2A4A')
    header_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    center = Alignment(horizontal='center', vertical='center')

    headers = ['product', 'chapter'] + LANG_CODES
    ws.append(headers)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # 欄位寬度
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    for ci in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 30

    ws.freeze_panes = 'C2'
    wb.save(output_path)


def get_column_letter(n: int) -> str:
    result = ''
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result

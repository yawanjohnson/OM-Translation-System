"""
idml_patcher.py - IDML 自動修正核心模組

功能：
  1. 在 IDML 的 Stories 中找到指定文字並替換
  2. 依狀態套用不同顏色標記，在 InDesign 中識別翻譯狀態
  3. 產生 Excel 比對報告（成功清單 + 未找到清單）

三色標記規則：
  - 黑色（無標記）：成功套用翻譯，正常顯示
  - 綠色（OM_GreenMark）：資料庫中找不到翻譯，保留英文原文
  - 紅色（OM_RedMark）：PM 確認後需更新的內容（後續修改用）

字串分割邏輯（Level 1）：
  若搜尋文字完整存在於單一 Content 元素中，則精確分割該
  CharacterStyleRange 為三段（前/替換/後），只有替換部分套色。

字串分割邏輯（Level 2）：
  若搜尋文字完整存在於一個 ParagraphStyleRange 的合併文字中，
  但跨越多個 CharacterStyleRange，則整段落重建（保留段落樣式，
  字元樣式簡化）。
"""

import re
import zipfile
import copy
import os
from io import BytesIO
from lxml import etree
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont
from modules.db_manager import split_prefix_suffix

# ------------------------------------------------------------------ #
# 常數
# ------------------------------------------------------------------ #
# 紅色：PM 後續更新標記
RED_COLOR_SELF  = 'Color/OM_Red'
RED_COLOR_NAME  = 'OM_Red'
RED_STYLE_SELF  = 'CharacterStyle/OM_RedMark'
RED_STYLE_NAME  = 'OM RedMark'

# 綠色：翻譯缺失（保留英文）標記
GREEN_COLOR_SELF = 'Color/OM_Green'
GREEN_COLOR_NAME = 'OM_Green'
GREEN_STYLE_SELF = 'CharacterStyle/OM_GreenMark'
GREEN_STYLE_NAME = 'OM GreenMark'

# 橘色：疑似錯字與相似句型標記
ORANGE_COLOR_SELF = 'Color/OM_Orange'
ORANGE_COLOR_NAME = 'OM_Orange'
ORANGE_STYLE_SELF = 'CharacterStyle/OM_OrangeMark'
ORANGE_STYLE_NAME = 'OM OrangeMark'

NO_CHAR_STYLE   = 'CharacterStyle/$ID/[No character style]'

LANG_CODES = [
    'ENG','GER','DUT','DAN','FRE','SPA','ITA','GRK',
    'POL','PRB','RUS','CHT','JPN','KOR','VTM','THI','ARB','TRK','CHS'
]


def _normalize_text(text: str) -> str:
    """比對前正規化：NBSP→空格、多餘空白合併、strip、轉小寫。
    僅用於比對，不改變寫回 XML 的內容。"""
    if not text:
        return ''
    # 將 \r, \n, \t 轉為空格，並移除 \x00-\x1f 控制字元
    text = re.sub(r'[\r\n\t]', ' ', text)
    text = re.sub(r'[\x00-\x1f]', '', text)
    # 轉換 NBSP, Unicode分行符和全形空白為普通空白
    text = (text.replace('\u00a0', ' ')
                .replace('\u2028', ' ')
                .replace('\u2029', ' ')
                .replace('\u3000', ' '))
    # 合併多個連續空格
    text = re.sub(r' +', ' ', text)
    return text.strip().lower()


def _remove_color_override(csr):
    """移除 CharacterStyleRange 的 FillColor 覆蓋，恢復段落樣式預設（黑色）。"""
    for child in list(csr):
        if _local(child.tag) == 'Properties':
            for prop_child in list(child):
                if _local(prop_child.tag) == 'FillColor':
                    child.remove(prop_child)
            break


def _apply_cjk_font_fix(csr):
    """將 CharacterStyleRange 的字體改為 Noto Sans CJK JP 且 FontStyle 設為 Regular，預防東亞字元缺字粉紅方塊。"""
    csr.set('AppliedFont', 'Noto Sans CJK JP')
    csr.set('FontStyle', 'Regular')




def _local(tag) -> str:
    if not isinstance(tag, str):
        return ''
    return tag.split('}')[-1] if '}' in tag else tag


# ------------------------------------------------------------------ #
# 主入口
# ------------------------------------------------------------------ #

def patch_idml(
    idml_path: str,
    instructions: list,
    output_idml_path: str,
    output_excel_path: str,
) -> dict:
    """
    修正 IDML 並產生報告。

    Args:
        idml_path:        來源 IDML 路徑
        instructions:     [{'lang_code','find','replace','note'}, ...]
        output_idml_path: 輸出標記版 IDML 路徑
        output_excel_path:輸出 Excel 報告路徑

    Returns:
        {'changes': [...], 'not_found': [...]}
    """
    # 讀取所有 IDML 內容
    with zipfile.ZipFile(idml_path, 'r') as zf:
        file_map = {name: zf.read(name) for name in zf.namelist()}
        zip_infos = zf.infolist()

    # 注入顏色樣式資源（紅色 + 綠色）
    file_map = _inject_color_styles(file_map)

    # 逐條執行修改指示
    all_changes = []
    not_found   = []
    story_names = sorted(n for n in file_map if n.startswith('Stories/'))

    # 排序：成功翻譯（黑色）先跑，翻譯缺失（綠色）後跑，避免順序污染
    ordered_instructions = sorted(
        instructions,
        key=lambda x: (1 if x.get('mark_green') else 0, 1 if x.get('mark_red') else 0)
    )

    for instr in ordered_instructions:
        find_text    = (instr.get('find') or '').strip()
        replace_text = (instr.get('replace') or '').strip()
        lang_code    = instr.get('lang_code', '')
        note         = instr.get('note', '')
        mark_red     = instr.get('mark_red', False)
        mark_green   = instr.get('mark_green', False)
        mark_orange  = instr.get('mark_orange', False)
        exact_match  = instr.get('exact_match', False)

        if not find_text:
            continue

        instr_changes = []

        for sname in story_names:
            new_xml, story_changes = _process_story(
                file_map[sname], find_text, replace_text, mark_red, mark_green, exact_match, lang_code, mark_orange
            )
            if story_changes:
                file_map[sname] = new_xml
                for sc in story_changes:
                    instr_changes.append({
                        'lang_code':  lang_code,
                        'story':      sname,
                        'find':       find_text,
                        'replace':    replace_text,
                        'note':       note,
                        'count':      sc.get('count', 1),
                        'mark_red':   mark_red,
                        'mark_green': mark_green,
                        'mark_orange': mark_orange,
                    })

        if instr_changes:
            all_changes.extend(instr_changes)
        else:
            not_found.append({
                'lang_code': lang_code,
                'find':      find_text,
                'replace':   replace_text,
                'note':      note,
            })

    # 寫出標記版 IDML
    _write_idml(file_map, zip_infos, output_idml_path)

    # 寫出 Excel 報告
    _write_excel_report(all_changes, not_found, output_excel_path)

    return {'changes': all_changes, 'not_found': not_found}


# ------------------------------------------------------------------ #
# 注入顏色樣式（紅色 + 綠色）
# ------------------------------------------------------------------ #

def _inject_color_styles(file_map: dict) -> dict:
    """在 Graphic.xml 及 Styles.xml 中注入紅色與綠色資源（冪等）。"""

    # --- Graphic.xml：新增顏色 ---
    if 'Resources/Graphic.xml' in file_map:
        tree = etree.fromstring(file_map['Resources/Graphic.xml'])

        # 紅色 CMYK(0,100,100,0)
        if tree.find(f'.//*[@Self="{RED_COLOR_SELF}"]') is None:
            color = etree.SubElement(tree, 'Color')
            color.set('Self',                RED_COLOR_SELF)
            color.set('Name',                RED_COLOR_NAME)
            color.set('Space',               'CMYK')
            color.set('ColorModel',          'Process')
            color.set('ColorValue',          '0 100 100 0')
            color.set('ColorOverride',       'Speciality')
            color.set('AlternateSpace',      'NoAlternateColorSpace')
            color.set('AlternateColorValue', '')
            color.set('Visible',             'true')
            color.set('ColorEditable',       'true')

        # 綠色 CMYK(75,0,100,0)
        if tree.find(f'.//*[@Self="{GREEN_COLOR_SELF}"]') is None:
            color = etree.SubElement(tree, 'Color')
            color.set('Self',                GREEN_COLOR_SELF)
            color.set('Name',                GREEN_COLOR_NAME)
            color.set('Space',               'CMYK')
            color.set('ColorModel',          'Process')
            color.set('ColorValue',          '75 0 100 0')
            color.set('ColorOverride',       'Speciality')
            color.set('AlternateSpace',      'NoAlternateColorSpace')
            color.set('AlternateColorValue', '')
            color.set('Visible',             'true')
            color.set('ColorEditable',       'true')

        # 橘色 CMYK(0,60,100,0)
        if tree.find(f'.//*[@Self="{ORANGE_COLOR_SELF}"]') is None:
            color = etree.SubElement(tree, 'Color')
            color.set('Self',                ORANGE_COLOR_SELF)
            color.set('Name',                ORANGE_COLOR_NAME)
            color.set('Space',               'CMYK')
            color.set('ColorModel',          'Process')
            color.set('ColorValue',          '0 60 100 0')
            color.set('ColorOverride',       'Speciality')
            color.set('AlternateSpace',      'NoAlternateColorSpace')
            color.set('AlternateColorValue', '')
            color.set('Visible',             'true')
            color.set('ColorEditable',       'true')

        file_map['Resources/Graphic.xml'] = _to_bytes(tree)

    # --- Styles.xml：新增字元樣式 ---
    if 'Resources/Styles.xml' in file_map:
        tree = etree.fromstring(file_map['Resources/Styles.xml'])
        csg = (
            tree.find('.//{http://ns.adobe.com/AdobeInDesign/idml/1.0/packaging}RootCharacterStyleGroup')
            or tree.find('.//RootCharacterStyleGroup')
        )
        if csg is not None:
            # 紅色字元樣式
            if csg.find(f'.//*[@Self="{RED_STYLE_SELF}"]') is None:
                cs = etree.SubElement(csg, 'CharacterStyle')
                cs.set('Self',      RED_STYLE_SELF)
                cs.set('Name',      RED_STYLE_NAME)
                cs.set('Imported',  'false')
                cs.set('FillColor', RED_COLOR_SELF)
                cs.set('FillTint',  '100')
                props = etree.SubElement(cs, 'Properties')
                bo = etree.SubElement(props, 'BasedOn')
                bo.set('type', 'string')
                bo.text = '$ID/[No character style]'

            # 綠色字元樣式
            if csg.find(f'.//*[@Self="{GREEN_STYLE_SELF}"]') is None:
                cs = etree.SubElement(csg, 'CharacterStyle')
                cs.set('Self',      GREEN_STYLE_SELF)
                cs.set('Name',      GREEN_STYLE_NAME)
                cs.set('Imported',  'false')
                cs.set('FillColor', GREEN_COLOR_SELF)
                cs.set('FillTint',  '100')
                props = etree.SubElement(cs, 'Properties')
                bo = etree.SubElement(props, 'BasedOn')
                bo.set('type', 'string')
                bo.text = '$ID/[No character style]'

            # 橘色字元樣式
            if csg.find(f'.//*[@Self="{ORANGE_STYLE_SELF}"]') is None:
                cs = etree.SubElement(csg, 'CharacterStyle')
                cs.set('Self',      ORANGE_STYLE_SELF)
                cs.set('Name',      ORANGE_STYLE_NAME)
                cs.set('Imported',  'false')
                cs.set('FillColor', ORANGE_COLOR_SELF)
                cs.set('FillTint',  '100')
                props = etree.SubElement(cs, 'Properties')
                bo = etree.SubElement(props, 'BasedOn')
                bo.set('type', 'string')
                bo.text = '$ID/[No character style]'

        file_map['Resources/Styles.xml'] = _to_bytes(tree)

    return file_map


# ------------------------------------------------------------------ #
# 處理 Story
# ------------------------------------------------------------------ #

def _process_story(
    xml_bytes: bytes,
    find: str,
    replace: str,
    mark_red: bool = False,
    mark_green: bool = False,
    exact_match: bool = False,
    lang_code: str = '',
    mark_orange: bool = False,
):
    """
    在 Story XML 中搜尋並替換文字。
    exact_match=True 時只在整個段落文字完全相符時才替換（正規化比對），
    防止子字串誤改其他段落。
    回傳 (new_xml_bytes, [change_info, ...])
    """
    try:
        tree = etree.fromstring(xml_bytes)
    except etree.XMLSyntaxError:
        return xml_bytes, []

    changes = []

    # 找到 Story 根元素
    story_elem = None
    for el in tree.iter():
        if _local(el.tag) == 'Story':
            story_elem = el
            break
    if story_elem is None:
        return xml_bytes, []

    # 逐段落處理
    for psr in list(story_elem.iter()):
        if _local(psr.tag) != 'ParagraphStyleRange':
            continue
        c = _replace_in_paragraph(psr, find, replace, mark_red, mark_green, exact_match, lang_code, mark_orange)
        if c:
            changes.append({'count': c})

    if not changes:
        return xml_bytes, []

    return _to_bytes(tree), changes


def _apply_color_override(csr, color_self: str):
    """在 CharacterStyleRange 中注入指定 FillColor 的本地樣式覆蓋，保留原字型不跑版/變亂碼。"""
    properties_el = None
    for child in csr:
        if _local(child.tag) == 'Properties':
            properties_el = child
            break
    if properties_el is None:
        properties_el = etree.SubElement(csr, 'Properties')

    fill_color_el = None
    for child in properties_el:
        if _local(child.tag) == 'FillColor':
            fill_color_el = child
            break
    if fill_color_el is None:
        fill_color_el = etree.SubElement(properties_el, 'FillColor')
        fill_color_el.set('type', 'string')
    fill_color_el.text = color_self

def _match_casing(original: str, replacement: str) -> str:
    """根據原文的大小寫特徵，自動將譯文進行大小寫對齊（只在原文為全大寫時將譯文轉全大寫，避免內文跑掉）。"""
    if not original or not replacement:
        return replacement
    # 僅在原文為全大寫且長度大於 1 時，才將譯文轉為全大寫（保留如標題大小寫等）
    if original.isupper() and len(original) > 1:
        return replacement.upper()
    return replacement


def _combine_replace_suffix(replace: str, suffix: str) -> tuple[str, str]:
    """避免翻譯後與原後置標點符號重複（例如 '翻譯。.' 或 '翻譯..' 變成重複點點）。"""
    if not suffix:
        return replace, suffix
    
    # 常用結尾標點符號
    ending_puncts = ('.', '。', '!', '！', '?', '？', ';', '；')
    
    # 如果譯文已經有結尾標點，則去除後置符號中多餘的結尾標點
    if replace and replace[-1] in ending_puncts:
        new_suffix = "".join([c for c in suffix if c not in ending_puncts])
        return replace, new_suffix
        
    return replace, suffix


def _drop_english_bracket_suffix(suffix: str, lang_code: str) -> str:
    """當目標語言不是英文時，若 suffix 是純英文多詞括號說明文字
    （例如 '(displayed when contact is made with both pulse grips)'），
    則自動清除，防止翻譯完成後英文括號殘留在譯文末尾。
    
    判斷邏輯：括號內容只含英文字母與空格（無數字/特殊碼）且超過 2 個單詞。
    這樣可以保留零件編號括號（如 '(21R)', '(30)'）不被誤刪。
    """
    if not suffix or not lang_code or lang_code == 'ENG':
        return suffix
    # 匹配純 ASCII 字母/空格的括號（可能前後有空白或句點）
    m = re.match(r'^(\s*\(([a-zA-Z][a-zA-Z\s]*)\)\s*[.\s]*)$', suffix)
    if m:
        bracket_content = m.group(2).strip()
        # 超過 2 個單詞才視為說明文字，單字縮寫或代碼不清除
        if len(bracket_content.split()) >= 3:
            return ''
    return suffix


def _split_csr_by_br(psr):
    """將 ParagraphStyleRange 中的 CharacterStyleRange 依照 Br (軟換行) 拆分為多個獨立的 CharacterStyleRange。
    這可以完全解決當同一段落有多行文字且共享同一個 CSR 時，其中一行缺失翻譯（綠字）導致整段文字全部變綠色的問題。
    """
    children = list(psr)
    for child in children:
        psr.remove(child)
        
    for child in children:
        if _local(child.tag) != 'CharacterStyleRange':
            psr.append(child)
            continue
            
        # 建立全新的 CSR 範本，清空子節點，但保留屬性與 Properties
        current_csr = copy.deepcopy(child)
        for sub in list(current_csr):
            current_csr.remove(sub)
            
        properties_el = None
        for sub in child:
            if _local(sub.tag) == 'Properties':
                properties_el = copy.deepcopy(sub)
                break
        if properties_el is not None:
            current_csr.append(properties_el)
            
        # 走訪原 CSR 內的所有子標籤進行拆分
        for sub in child:
            if _local(sub.tag) == 'Properties':
                continue
            elif _local(sub.tag) == 'Br':
                # 當遇到 Br 時，如果目前的 CSR 已有實質文字內容，先將它 append 進 Paragraph
                if len([c for c in current_csr if _local(c.tag) != 'Properties']) > 0:
                    psr.append(current_csr)
                
                # 將 Br 本身作為 Paragraph 的直接子節點 (即 CharacterStyleRange 的 Sibling)
                psr.append(sub)
                
                # 重新開一個乾淨的 CSR 繼承原樣式
                current_csr = copy.deepcopy(child)
                for c in list(current_csr):
                    current_csr.remove(c)
                if properties_el is not None:
                    current_csr.append(copy.deepcopy(properties_el))
            else:
                current_csr.append(sub)
                
        # 寫入最後一個剩餘的 CSR
        if len([c for c in current_csr if _local(c.tag) != 'Properties']) > 0:
            psr.append(current_csr)


def _replace_in_paragraph(
    psr,
    find: str,
    replace: str,
    mark_red: bool = False,
    mark_green: bool = False,
    exact_match: bool = False,
    lang_code: str = '',
    mark_orange: bool = False,
) -> int:
    """
    在一個 ParagraphStyleRange 中執行非破壞性的搜尋與替換。
    100% 保留所有原本字元樣式（CharacterStyleRange）與各國語言字型設定。
    """
    # 0. 執行 CSR 軟換行拆分，預防各行顏色互相污染
    _split_csr_by_br(psr)

    csrs = [c for c in psr if _local(c.tag) == 'CharacterStyleRange']

    # 1. 建立字元映射表 (走訪 psr 所有子節點，包含 Br，確保與 idml_parser.py 讀出的一致)
    char_map = []
    for child in psr:
        local = _local(child.tag)
        if local == 'CharacterStyleRange':
            for gchild in child:
                glocal = _local(gchild.tag)
                if glocal == 'Content':
                    text = gchild.text or ''
                    for i, char in enumerate(text):
                        mapped_char = '\n' if char in ('\r', '\u2028', '\u2029') else char
                        char_map.append({
                            'char': mapped_char,
                            'csr': child,
                            'content': gchild,
                            'index': i
                        })
                elif glocal == 'Br':
                    char_map.append({
                        'char': '\n',
                        'csr': child,
                        'content': gchild,
                        'index': None
                    })
        elif local == 'Br':
            char_map.append({
                'char': '\n',
                'csr': None,
                'content': child,
                'index': None
            })

    full_text = "".join([item['char'] for item in char_map])

    # --- exact_match 模式 ---
    if exact_match:
        lines = full_text.split('\n')
        current_pos = 0
        matched_any = False
        marked_csrs = set()

        for line in lines:
            line_start = full_text.find(line, current_pos)
            if line_start != -1:
                # 拆分前置與後置符號並進行正規化比對
                prefix, core, suffix = split_prefix_suffix(line)
                if _normalize_text(core) == _normalize_text(find):
                    core_start = line_start + len(prefix)
                    core_end = core_start + len(core)
                    line_end = line_start + len(line)

                    # 收集受影響之 CharacterStyleRange
                    for item in char_map[core_start:core_end]:
                        if item['csr'] is not None:
                            marked_csrs.add(item['csr'])

                    # 決定置換後的核心文字與後綴，避免重複標點
                    if mark_green:
                        actual_replace = core
                        new_suffix = suffix
                    else:
                        actual_replace = _match_casing(core, replace)
                        actual_replace, new_suffix = _combine_replace_suffix(actual_replace, suffix)
                        # Fix: 自動清除純英文多詞括號說明後綴，防止翻譯後英文括號殘留
                        new_suffix = _drop_english_bracket_suffix(new_suffix, lang_code)

                    # 置換整行文字，清空 core 及原後綴，保留 prefix
                    content_chars = {}
                    for item in char_map:
                        if item['index'] is not None:
                            c = item['content']
                            if c not in content_chars:
                                content_chars[c] = list(c.text or '')

                    # Check if colons exist in both core and actual_replace to preserve title/body split
                    orig_colon_pos = core.find(':')
                    if orig_colon_pos == -1:
                        orig_colon_pos = core.find('：')
                    
                    repl_colon_pos = actual_replace.find(':')
                    if repl_colon_pos == -1:
                        repl_colon_pos = actual_replace.find('：')

                    split_success = False
                    if orig_colon_pos != -1 and repl_colon_pos != -1:
                        core_offset = len(prefix)
                        orig_title_len = orig_colon_pos + 1
                        body_start_idx = line_start + core_offset + orig_title_len
                        
                        first_item = char_map[line_start]
                        next_style_idx = body_start_idx
                        while next_style_idx < line_end and next_style_idx < len(char_map):
                            if char_map[next_style_idx]['content'] != first_item['content']:
                                break
                            next_style_idx += 1
                        
                        if next_style_idx < line_end and next_style_idx < len(char_map):
                            body_start_idx = next_style_idx
                            
                            replace_title = actual_replace[:repl_colon_pos + 1]
                            replace_body = actual_replace[repl_colon_pos + 1:]
                            
                            new_title_text = prefix + replace_title
                            new_body_text = replace_body + new_suffix
                            
                            body_first_item = char_map[body_start_idx]
                            
                            if first_item['index'] is not None and body_first_item['index'] is not None:
                                content_chars[first_item['content']][first_item['index']] = new_title_text
                                for item in char_map[line_start + 1 : body_start_idx]:
                                    if item['index'] is not None:
                                        content_chars[item['content']][item['index']] = ""
                                
                                content_chars[body_first_item['content']][body_first_item['index']] = new_body_text
                                for item in char_map[body_start_idx + 1 : line_end]:
                                    if item['index'] is not None:
                                        content_chars[item['content']][item['index']] = ""
                                
                                split_success = True

                    if not split_success:
                        new_line_text = prefix + actual_replace + new_suffix
                        first_item = char_map[line_start]
                        if first_item['index'] is not None:
                            content_chars[first_item['content']][first_item['index']] = new_line_text
                        for item in char_map[line_start + 1 : line_end]:
                            if item['index'] is not None:
                                content_chars[item['content']][item['index']] = ""

                    for content, chars in content_chars.items():
                        content.text = "".join(chars)

                    matched_any = True
                    break
                current_pos = line_start + len(line) + 1

        if matched_any:
            if marked_csrs:
                if mark_red:
                    for csr in marked_csrs:
                        _apply_color_override(csr, RED_COLOR_SELF)
                elif mark_green:
                    for csr in marked_csrs:
                        _apply_color_override(csr, GREEN_COLOR_SELF)
                elif mark_orange:
                    for csr in marked_csrs:
                        _apply_color_override(csr, ORANGE_COLOR_SELF)
                else:
                    for csr in marked_csrs:
                        _remove_color_override(csr)
                    if lang_code in ('CHT', 'CHS', 'JPN', 'KOR'):
                        for csr in marked_csrs:
                            _apply_cjk_font_fix(csr)
            return 1
        return 0

    # --- 非 exact_match 模式 (子字串模糊匹配，支援定位點與大小寫容錯) ---
    # 1. 產生無空白的 find 字串
    norm_find = _normalize_text(find)
    spaceless_find = re.sub(r'\s+', '', norm_find)
    if not spaceless_find:
        return 0

    # 2. 建立無空白的 full_text 映射表，對應回原始 char_map 中的索引
    spaceless_full = ""
    orig_indices = []
    for idx, item in enumerate(char_map):
        char = item['char']
        # 忽略所有換行、定位點與空白控制字元
        if char.isspace() or ord(char) < 32 or char in ('\xa0', '\u2002', '\u2003', '\u2009', '\u2028', '\u2029', '\u3000'):
            continue
        spaceless_full += char.lower()
        orig_indices.append(idx)

    # 3. 搜尋匹配區間
    match_indices = []  # 儲存 (orig_start, orig_end) 區間元組
    start = 0
    while True:
        idx = spaceless_full.find(spaceless_find, start)
        if idx == -1:
            break
        orig_start = orig_indices[idx]
        orig_end = orig_indices[idx + len(spaceless_find) - 1] + 1
        match_indices.append((orig_start, orig_end))
        start = idx + len(spaceless_find)

    if not match_indices:
        return 0

    content_chars = {}
    for item in char_map:
        if item['index'] is not None:
            content = item['content']
            if content not in content_chars:
                content_chars[content] = {}
            content_chars[content][item['index']] = item['char']

    marked_csrs = set()
    for orig_start, orig_end in match_indices:
        match_items = char_map[orig_start : orig_end]
        first_item = match_items[0]
        first_content = first_item['content']
        first_idx = first_item['index']
        effective_find = "".join([item['char'] for item in match_items])
        idx = orig_start

        # Check if colons exist in both effective_find and replace to preserve title/body split
        orig_colon_pos = effective_find.find(':')
        if orig_colon_pos == -1:
            orig_colon_pos = effective_find.find('：')
        
        repl_colon_pos = replace.find(':')
        if repl_colon_pos == -1:
            repl_colon_pos = replace.find('：')

        split_success = False
        if orig_colon_pos != -1 and repl_colon_pos != -1:
            orig_title_len = orig_colon_pos + 1
            body_start_idx = idx + orig_title_len
            
            next_style_idx = body_start_idx
            while next_style_idx < idx + len(effective_find) and next_style_idx < len(char_map):
                if char_map[next_style_idx]['content'] != first_content:
                    break
                next_style_idx += 1
                
            if next_style_idx < idx + len(effective_find) and next_style_idx < len(char_map):
                body_start_idx = next_style_idx
                replace_title = replace[:repl_colon_pos + 1]
                replace_body = replace[repl_colon_pos + 1:]
                
                body_first_item = char_map[body_start_idx]
                
                if first_idx is not None and body_first_item['index'] is not None:
                    content_chars[first_content][first_idx] = replace_title
                    for item in match_items[1 : body_start_idx - idx]:
                        if item['index'] is not None:
                            content_chars[item['content']][item['index']] = ""
                    
                    content_chars[body_first_item['content']][body_first_item['index']] = replace_body
                    for item in match_items[body_start_idx - idx + 1 :]:
                        if item['index'] is not None:
                            content_chars[item['content']][item['index']] = ""
                    
                    if first_item['csr'] is not None:
                        marked_csrs.add(first_item['csr'])
                    if body_first_item['csr'] is not None:
                        marked_csrs.add(body_first_item['csr'])
                        
                    split_success = True

        if not split_success:
            if first_idx is not None:
                content_chars[first_content][first_idx] = replace
                if first_item['csr'] is not None:
                    marked_csrs.add(first_item['csr'])

            for item in match_items[1:]:
                if item['index'] is not None:
                    content_chars[item['content']][item['index']] = ""

    for content, char_dict in content_chars.items():
        sorted_indices = sorted(char_dict.keys())
        new_text = "".join([char_dict[i] for i in sorted_indices])
        content.text = new_text

    if marked_csrs:
        if mark_red:
            for csr in marked_csrs:
                _apply_color_override(csr, RED_COLOR_SELF)
        elif mark_green:
            for csr in marked_csrs:
                _apply_color_override(csr, GREEN_COLOR_SELF)
        else:
            for csr in marked_csrs:
                _remove_color_override(csr)
            if lang_code in ('CHT', 'CHS', 'JPN', 'KOR'):
                for csr in marked_csrs:
                    _apply_cjk_font_fix(csr)

    return len(match_indices)
def _para_full_text(psr) -> str:
    """萃取段落全文。"""
    parts = []
    for el in psr.iter():
        local = _local(el.tag)
        if local == 'Content' and el.text:
            parts.append(el.text)
        elif local == 'Br':
            parts.append('\n')
    return ''.join(parts)


# ------------------------------------------------------------------ #
# 寫出 IDML
# ------------------------------------------------------------------ #

def _write_idml(file_map: dict, zip_infos, output_path: str):
    """將修改後的 IDML 寫出為 ZIP（保留 mimetype 不壓縮）。"""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with zipfile.ZipFile(output_path, 'w') as zout:
        for info in zip_infos:
            content = file_map.get(info.filename)
            if content is None:
                continue
            compress = zipfile.ZIP_STORED if info.filename == 'mimetype' else zipfile.ZIP_DEFLATED
            zout.writestr(info.filename, content, compress)


# ------------------------------------------------------------------ #
# Excel 報告
# ------------------------------------------------------------------ #

def get_indesign_rich_diff(str1: str, str2: str) -> CellRichText:
    """
    InDesign 英文原文：只將 InDesign 中多打的字/錯字標記為紅色粗體，其餘正常顯示。
    """
    red_font = InlineFont(color='FF0000', rFont='Arial', sz=10, b=True)
    norm_font = InlineFont(rFont='Arial', sz=10)
    
    import difflib
    s = difflib.SequenceMatcher(None, str1, str2)
    rt = CellRichText()
    
    for tag, i1, i2, j1, j2 in s.get_opcodes():
        if tag == 'equal':
            rt.append(TextBlock(norm_font, str1[i1:i2]))
        elif tag == 'replace':
            rt.append(TextBlock(red_font, str1[i1:i2]))
        elif tag == 'delete':
            rt.append(TextBlock(red_font, str1[i1:i2]))
            
    return rt


def get_database_rich_diff(str1: str, str2: str) -> CellRichText:
    """
    資料庫最相似原文：只將資料庫中多出/正確的字標記為綠色粗體，其餘正常顯示。
    """
    green_font = InlineFont(color='008000', rFont='Arial', sz=10, b=True)
    norm_font = InlineFont(rFont='Arial', sz=10)
    
    import difflib
    s = difflib.SequenceMatcher(None, str1, str2)
    rt = CellRichText()
    
    for tag, i1, i2, j1, j2 in s.get_opcodes():
        if tag == 'equal':
            rt.append(TextBlock(norm_font, str2[j1:j2]))
        elif tag == 'replace':
            rt.append(TextBlock(green_font, str2[j1:j2]))
        elif tag == 'insert':
            rt.append(TextBlock(green_font, str2[j1:j2]))
            
    return rt


def _write_excel_report(changes: list, not_found: list, output_path: str):
    """
    Excel report (4 sheets):
      Sheet 1: Translations applied
      Sheet 2: Missing translations (kept English, green in IDML)
      Sheet 3: Warnings for typos / similar sentences (kept English, orange in IDML)
      Sheet 4: Text not found in IDML at all
    All content cells use black font. Color is only on headers.
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = openpyxl.Workbook()

    NORM_FONT = Font(name='Arial', size=10)
    BOLD_FONT = Font(name='Arial', size=10, bold=True)
    ALT_FILL  = PatternFill('solid', fgColor='F4F6FB')
    CENTER    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT      = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    def _make_header(ws, headers, col_widths, hdr_color='1B2A4A'):
        hdr_fill = PatternFill('solid', fgColor=hdr_color)
        hdr_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        ws.row_dimensions[1].height = 22
        ws.append(headers)
        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(1, ci)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = CENTER
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = 'A2'
        ws.sheet_view.showGridLines = False

    def _set_row(ws, row_idx, col_count, is_alt):
        if is_alt:
            for ci in range(1, col_count + 1):
                ws.cell(row_idx, ci).fill = ALT_FILL
        for ci in range(1, col_count + 1):
            ws.cell(row_idx, ci).font = NORM_FONT
            ws.cell(row_idx, ci).alignment = LEFT
        ws.cell(row_idx, 1).alignment = CENTER
        ws.cell(row_idx, 2).alignment = CENTER

    applied  = [c for c in changes if not c.get('mark_green') and not c.get('mark_red') and not c.get('mark_orange')]
    applied  += [c for c in changes if c.get('mark_red')]
    missing  = [c for c in changes if c.get('mark_green')]
    warnings = [c for c in changes if c.get('mark_orange')]

    # Sheet 1
    ws1 = wb.active
    ws1.title = '✅ 翻譯已套用'
    _make_header(ws1,
        ['#', '語言', '英文原文', '套用後文字', '備註', 'Story 位置'],
        [5, 8, 48, 48, 20, 35],
        '1B6B3A')
    for i, ch in enumerate(applied, 1):
        row_idx = i + 1
        ws1.row_dimensions[row_idx].height = 36
        ws1.append([i, ch.get('lang_code',''), ch.get('find',''), ch.get('replace',''),
                    ch.get('note',''), ch.get('story','').replace('Stories/','')])
        _set_row(ws1, row_idx, 6, i % 2 == 0)
    ws1.append([])
    ws1.append(['', f'共 {len(applied)} 條翻譯已套用'])
    ws1.cell(ws1.max_row, 2).font = BOLD_FONT

    # Sheet 2
    ws2 = wb.create_sheet('🟢 翻譯缺失')
    _make_header(ws2,
        ['#', '語言', '英文原文（保留）', '資料庫狀態', '備註', 'Story 位置'],
        [5, 8, 55, 20, 20, 35],
        '4A7A1A')
    for i, ch in enumerate(missing, 1):
        row_idx = i + 1
        ws2.row_dimensions[row_idx].height = 36
        ws2.append([i, ch.get('lang_code',''), ch.get('find',''), '資料庫無此翻譯',
                    ch.get('note',''), ch.get('story','').replace('Stories/','')])
        _set_row(ws2, row_idx, 6, i % 2 == 0)
    ws2.append([])
    if missing:
        ws2.append(['', f'共 {len(missing)} 條翻譯缺失，IDML 中以綠色標記'])
    else:
        ws2.append(['', '🎉 無翻譯缺失，所有文字均已套用！'])
    ws2.cell(ws2.max_row, 2).font = BOLD_FONT

# Sheet 3 (疑似錯字與相似句型警告)
    ws_warn = wb.create_sheet('⚠️ 疑似錯字與相似句型')
    _make_header(ws_warn,
        ['#', '語言', '英文原文（疑似錯字）', '預計套用翻譯 / 備註', '資料庫最相似原文', 'Story 位置'],
        [5, 8, 55, 48, 55, 35],
        'D27D2D') # 橘色標記
    for i, ch in enumerate(warnings, 1):
        row_idx = i + 1
        ws_warn.row_dimensions[row_idx].height = 36
        # From note format: "DB Similar Found (ID: 812, Similarity: 99%): <db_eng_text>"
        note = ch.get('note', '')
        db_text = ''
        if '): ' in note:
            parts = note.split('): ', 1)
            note_prefix = parts[0] + ')'
            db_text = parts[1]
        else:
            note_prefix = note
        
        rich_find = get_indesign_rich_diff(ch.get('find',''), db_text)
        rich_db = get_database_rich_diff(ch.get('find',''), db_text)
        
        ws_warn.append([i, ch.get('lang_code',''), "",
                    note_prefix, "", ch.get('story','').replace('Stories/','')])
        ws_warn.cell(row_idx, 3).value = rich_find
        ws_warn.cell(row_idx, 5).value = rich_db
        _set_row(ws_warn, row_idx, 6, i % 2 == 0)
    ws_warn.append([])
    if warnings:
        ws_warn.append(['', f'共 {len(warnings)} 條疑似錯字或高度相似文字，IDML 中以橘色標記'])
    else:
        ws_warn.append(['', '🎉 未偵測到疑似錯字或高度相似之未翻譯句型。'])
    ws_warn.cell(ws_warn.max_row, 2).font = BOLD_FONT

    # Sheet 4
    ws3 = wb.create_sheet('❌ IDML找不到')
    _make_header(ws3,
        ['#', '語言', '搜尋文字（IDML中無此段落）', '預計套用翻譯', '備註'],
        [5, 8, 55, 48, 20],
        '7B1A1A')
    for i, nf in enumerate(not_found, 1):
        row_idx = i + 1
        ws3.row_dimensions[row_idx].height = 36
        ws3.append([i, nf.get('lang_code',''), nf.get('find',''),
                    nf.get('replace',''), nf.get('note','')])
        _set_row(ws3, row_idx, 5, i % 2 == 0)
    ws3.append([])
    if not_found:
        ws3.append(['', f'共 {len(not_found)} 條文字在 IDML 中找不到對應段落'])
    else:
        ws3.append(['', '🎉 IDML 中所有段落均已找到！'])
    # Sheet 1
    ws1 = wb.active
    ws1.title = '✅ 翻譯已套用'
    _make_header(ws1,
        ['#', '語言', '英文原文', '套用後文字', '備註', 'Story 位置'],
        [5, 8, 48, 48, 20, 35],
        '1B6B3A')
    for i, ch in enumerate(applied, 1):
        row_idx = i + 1
        ws1.row_dimensions[row_idx].height = 36
        ws1.append([i, ch.get('lang_code',''), ch.get('find',''), ch.get('replace',''),
                    ch.get('note',''), ch.get('story','').replace('Stories/','')])
        _set_row(ws1, row_idx, 6, i % 2 == 0)
    ws1.append([])
    ws1.append(['', f'共 {len(applied)} 條翻譯已套用'])
    ws1.cell(ws1.max_row, 2).font = BOLD_FONT

    # Sheet 2
    ws2 = wb.create_sheet('🟢 翻譯缺失')
    _make_header(ws2,
        ['#', '語言', '英文原文（保留）', '資料庫狀態', '備註', 'Story 位置'],
        [5, 8, 55, 20, 20, 35],
        '4A7A1A')
    for i, ch in enumerate(missing, 1):
        row_idx = i + 1
        ws2.row_dimensions[row_idx].height = 36
        ws2.append([i, ch.get('lang_code',''), ch.get('find',''), '資料庫無此翻譯',
                    ch.get('note',''), ch.get('story','').replace('Stories/','')])
        _set_row(ws2, row_idx, 6, i % 2 == 0)
    ws2.append([])
    if missing:
        ws2.append(['', f'共 {len(missing)} 條翻譯缺失，IDML 中以綠色標記'])
    else:
        ws2.append(['', '🎉 無翻譯缺失，所有文字均已套用！'])
    ws2.cell(ws2.max_row, 2).font = BOLD_FONT

    # Sheet 3
    ws3 = wb.create_sheet('❌ IDML找不到')
    _make_header(ws3,
        ['#', '語言', '搜尋文字（IDML中無此段落）', '預計套用翻譯', '備註'],
        [5, 8, 55, 48, 20],
        '7B1A1A')
    for i, nf in enumerate(not_found, 1):
        row_idx = i + 1
        ws3.row_dimensions[row_idx].height = 36
        ws3.append([i, nf.get('lang_code',''), nf.get('find',''),
                    nf.get('replace',''), nf.get('note','')])
        _set_row(ws3, row_idx, 5, i % 2 == 0)
    ws3.append([])
    if not_found:
        ws3.append(['', f'共 {len(not_found)} 條文字在 IDML 中找不到對應段落'])
    else:
        ws3.append(['', '🎉 IDML 中所有段落均已找到！'])
    ws3.cell(ws3.max_row, 2).font = BOLD_FONT

    wb.save(output_path)



# ------------------------------------------------------------------ #
# 工具函式
# ------------------------------------------------------------------ #

def _to_bytes(tree) -> bytes:
    return etree.tostring(tree, xml_declaration=True, encoding='UTF-8', standalone=True)

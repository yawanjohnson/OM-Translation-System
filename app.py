"""
app.py - OM 多語言管理系統 Flask 後端
"""
import os
import uuid
import json
from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024  # 200 MB

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR  = os.path.join(BASE_DIR, 'uploads')
OUTPUT_DIR  = os.path.join(BASE_DIR, 'outputs')
DATA_DIR    = os.path.join(BASE_DIR, 'data')
DB_PATH     = os.path.join(DATA_DIR, 'translations.db')

for d in (UPLOAD_DIR, OUTPUT_DIR, DATA_DIR):
    os.makedirs(d, exist_ok=True)

from modules.db_manager import DBManager, LANG_CODES, LANG_NAMES, should_skip_translation, clean_text, split_prefix_suffix
from modules.idml_patcher import patch_idml
from modules.idml_parser import extract_stories, get_idml_info, extract_layout
from modules.importers.excel_importer import import_excel, generate_template
from modules.importers.idml_importer import preview_extract, extract_for_import
from modules.pm_review_generator import generate_pm_review_html
from modules.indesign_exporter import export_pdf_via_indesign

db = DBManager(DB_PATH)

ALLOWED_IDML  = {'idml'}
ALLOWED_EXCEL = {'xlsx', 'xls', 'csv'}
ALLOWED_INSTR = {'xlsx', 'xls', 'csv'}


def allowed(filename, exts):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in exts


def save_upload(file, exts) -> str:
    if not allowed(file.filename, exts):
        raise ValueError(f'不支援的檔案格式：{file.filename}')
    base, ext = os.path.splitext(file.filename)
    sec_base = secure_filename(base)
    if not sec_base:
        sec_base = 'file'
    name = f'{uuid.uuid4().hex}_{sec_base}{ext.lower()}'
    path = os.path.join(UPLOAD_DIR, name)
    file.save(path)
    return path


# ──────────────────────────────────────────
# 頁面
# ──────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


# ──────────────────────────────────────────
# 語言資訊
# ──────────────────────────────────────────

@app.route('/api/lang-info')
def api_lang_info():
    return jsonify({'codes': LANG_CODES, 'names': LANG_NAMES})


# ──────────────────────────────────────────
# 多語言資料庫 CRUD
# ──────────────────────────────────────────

@app.route('/api/translations', methods=['GET'])
def api_list():
    q        = request.args.get('q', '')
    lang     = request.args.get('lang', 'ENG')
    product  = request.args.get('product', '')
    chapter  = request.args.get('chapter', '')
    page     = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 50))
    result   = db.search(q, lang, product, chapter, page, per_page)
    return jsonify(result)


@app.route('/api/translations', methods=['POST'])
def api_add():
    data = request.json or {}
    try:
        tid = db.add(data)
        return jsonify({'ok': True, 'id': tid})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400


@app.route('/api/translations/<int:tid>', methods=['GET'])
def api_get(tid):
    row = db.get(tid)
    if row:
        return jsonify(row)
    return jsonify({'error': 'not found'}), 404


@app.route('/api/translations/<int:tid>', methods=['PUT'])
def api_update(tid):
    data = request.json or {}
    ok = db.update(tid, data)
    return jsonify({'ok': ok})


@app.route('/api/translations/<int:tid>', methods=['DELETE'])
def api_delete(tid):
    ok = db.delete(tid)
    return jsonify({'ok': ok})


@app.route('/api/translations/stats')
def api_stats():
    return jsonify(db.get_stats())


@app.route('/api/translations/history', methods=['GET'])
def api_get_history():
    """取得資料庫修改歷程。"""
    try:
        page = request.args.get('page', 1, type=int)
        tid = request.args.get('translation_id', None, type=int)
        per_page = 20
        result = db.get_history(page=page, page_size=per_page, translation_id=tid)
        return jsonify(result)
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/translations/history/<int:hid>/revert', methods=['POST'])
def api_revert_history(hid):
    """將資料庫內容回復至指定的歷史狀態。"""
    try:
        ok = db.revert_history(hid)
        if ok:
            return jsonify({'ok': True, 'message': '已成功回復資料至指定版本。'})
        else:
            return jsonify({'ok': False, 'message': '回復操作無效或資料未改變。'}), 400
    except Exception as e:
        return jsonify({'ok': False, 'message': f'回復失敗：{str(e)}'}), 500


@app.route('/api/translations/duplicates', methods=['GET'])
def api_get_duplicates():
    """掃描資料庫中的重複原文 (ENG) 條目。"""
    try:
        res = db.find_duplicate_groups()
        return jsonify({'ok': True, **res})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/translations/merge-duplicates', methods=['POST'])
def api_merge_duplicates():
    """執行合併重複原文。"""
    data = request.json or {}
    resolutions = data.get('resolutions', [])
    try:
        res = db.merge_duplicate_groups(resolutions)
        return jsonify(res)
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ──────────────────────────────────────────
# 匯入
# ──────────────────────────────────────────

@app.route('/api/import/excel', methods=['POST'])
def api_import_excel():
    """匯入 Excel/CSV 翻譯對照表到資料庫。"""
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': '請上傳檔案'}), 400
    try:
        path  = save_upload(request.files['file'], ALLOWED_EXCEL)
        rows  = import_excel(path)
        stats = db.bulk_upsert(rows)
        os.remove(path)
        return jsonify({'ok': True, **stats, 'total_rows': len(rows)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/import/excel-with-conflicts', methods=['POST'])
def api_import_excel_with_conflicts():
    """
    匯入 Excel/CSV 對照表，執行衝突檢測分流：
    - 無衝突的資料直接寫入/更新資料庫。
    - 有衝突的資料（英文相同但翻譯不同）寫入 pending_conflicts 待處理表。
    """
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': '請上傳檔案'}), 400
    try:
        path = save_upload(request.files['file'], ALLOWED_EXCEL)
        rows = import_excel(path)
        os.remove(path)
        
        # 檢查是否有 ENG 欄位
        has_eng_column = any('ENG' in r for r in rows)
        if rows and not has_eng_column:
            return jsonify({
                'ok': False, 
                'error': '檔案中未偵測到「英文 (ENG)」欄位，無法進行翻譯對照。請確認第一列欄位名稱。'
            }), 400

        imported_count = 0
        conflict_count = 0
        skipped_count = 0
        
        for data in rows:
            eng_val = data.get('ENG', '').strip() if 'ENG' in data and data['ENG'] is not None else ''
            if not eng_val:
                skipped_count += 1
                continue
                
            product = data.get('product', '') or ''
            chapter = data.get('chapter', '') or ''
            
            existing = db.lookup_eng(eng_val)
            if not existing:
                # 全新條目：直接寫入 translations 表
                db.add(data)
                imported_count += 1
            else:
                # 存在舊條目：比對各國語言翻譯是否衝突
                has_conflict_on_row = False
                conflicting_langs = {}
                non_conflicting_updates = {}
                
                for code in LANG_CODES:
                    if code == 'ENG':
                        continue
                    excel_translation = data.get(code, '')
                    if excel_translation is not None:
                        excel_translation = str(excel_translation).strip()
                    else:
                        excel_translation = ''
                        
                    db_translation = existing.get(code, '') or ''
                    
                    if excel_translation and db_translation:
                        # 兩者皆有翻譯，但字串不相同 -> 衝突！
                        if excel_translation != db_translation:
                            has_conflict_on_row = True
                            conflicting_langs[code] = (db_translation, excel_translation)
                    elif excel_translation and not db_translation:
                        # Excel 有新翻譯，資料庫是空的 -> 可以直接補齊，不視為衝突
                        non_conflicting_updates[code] = excel_translation
                
                if has_conflict_on_row:
                    # 寫入待處理衝突表
                    for lang, (db_val, import_val) in conflicting_langs.items():
                        db.add_pending_conflict(
                            product=product or existing.get('product', ''),
                            chapter=chapter or existing.get('chapter', ''),
                            eng_text=eng_val,
                            lang_code=lang,
                            db_val=db_val,
                            import_val=import_val
                        )
                        conflict_count += 1
                    # 對於此列中無衝突的新翻譯（例如其他語言有補齊的），我們先寫入 translations 表
                    if non_conflicting_updates:
                        db.update(existing['id'], non_conflicting_updates)
                else:
                    # 無任何語言翻譯衝突，直接更新 translations（包含補齊空白的翻譯，或 product/chapter 更新）
                    update_data = {}
                    if product: update_data['product'] = product
                    if chapter: update_data['chapter'] = chapter
                    for code, val in non_conflicting_updates.items():
                        update_data[code] = val
                        
                    if update_data:
                        db.update(existing['id'], update_data)
                    imported_count += 1
                    
        return jsonify({
            'ok': True,
            'imported_count': imported_count,
            'conflict_count': conflict_count,
            'skipped_count': skipped_count,
            'message': f'{imported_count} 筆資料已成功處理；{conflict_count} 個語言翻譯衝突已移至待處理佇列。'
        })
    except Exception as e:
        import traceback
        return jsonify({'ok': False, 'error': str(e), 'detail': traceback.format_exc()}), 500


@app.route('/api/import/pending-conflicts', methods=['GET'])
def api_get_pending_conflicts():
    """獲取待處理衝突清單"""
    try:
        conflicts = db.get_pending_conflicts()
        return jsonify({'ok': True, 'conflicts': conflicts, 'count': len(conflicts)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/import/resolve-conflicts', methods=['POST'])
def api_resolve_conflicts():
    """
    提交衝突解決決策：
    Body: {
        "resolutions": [
            { "pending_id": 12, "decision": "KEEP_IMPORT" },
            { "pending_id": 13, "decision": "KEEP_DB" }
        ]
    }
    """
    data = request.json or {}
    resolutions = data.get('resolutions', [])
    if not resolutions:
        return jsonify({'ok': False, 'error': '未提供任何解決方案'}), 400
        
    try:
        resolved_count = 0
        batch_id = uuid.uuid4().hex[:8]
        
        for res in resolutions:
            pid = res.get('pending_id')
            decision = res.get('decision') # 'KEEP_IMPORT' or 'KEEP_DB'
            
            if not pid or not decision:
                continue
                
            conflict = db.get_pending_conflict(pid)
            if not conflict:
                continue
                
            eng_text = conflict['eng_text']
            lang_code = conflict['lang_code']
            db_val = conflict['db_val']
            import_val = conflict['import_val']
            
            chosen_val = import_val if decision in ('KEEP_IMPORT', 'ADD_NEW') else db_val
            if decision == 'DELETE':
                chosen_val = '[已刪除]'
            
            # 1. 如果選擇匯入新版，更新 translations 表中對應翻譯
            if decision == 'KEEP_IMPORT':
                existing = db.lookup_eng(eng_text)
                if existing:
                    db.update(existing['id'], {lang_code: chosen_val})
            
            # 1b. 如果選擇新增，則在 translations 表中新增一行
            elif decision == 'ADD_NEW':
                db.add({
                    'product': conflict.get('product') or '',
                    'chapter': conflict.get('chapter') or '',
                    'ENG': eng_text,
                    lang_code: import_val
                })
            
            # 1c. 如果選擇刪除，則自 translations 資料表刪除該翻譯條目
            elif decision == 'DELETE':
                existing = db.lookup_eng(eng_text)
                if existing:
                    db.delete(existing['id'])
                    
            # 2. 寫入衝突日誌表
            db.add_conflict_log(
                batch_id=batch_id,
                eng_text=eng_text,
                lang_code=lang_code,
                db_val=db_val,
                import_val=import_val,
                chosen_val=chosen_val,
                decision=decision,
                suppress_export=True
            )
            
            # 3. 自 pending_conflicts 刪除
            db.delete_pending_conflict(pid)
            resolved_count += 1
            
        if resolved_count > 0:
            db.export_conflict_logs_to_git_json()
            
        return jsonify({'ok': True, 'resolved_count': resolved_count, 'batch_id': batch_id})
    except Exception as e:
        import traceback
        return jsonify({'ok': False, 'error': str(e), 'detail': traceback.format_exc()}), 500


@app.route('/api/import/conflict-logs', methods=['GET'])
def api_get_conflict_logs():
    """獲取已解決衝突的歷史紀錄"""
    q = request.args.get('q', '')
    lang = request.args.get('lang', '')
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 50))
    try:
        result = db.get_conflict_logs(query=q, lang_code=lang, page=page, per_page=per_page)
        return jsonify({'ok': True, **result})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/import/idml-preview', methods=['POST'])
def api_import_idml_preview():
    """預覽 IDML 萃取結果（尚未存入資料庫）。"""
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': '請上傳 IDML'}), 400
    lang = request.form.get('lang_code', 'ENG')
    try:
        path   = save_upload(request.files['file'], ALLOWED_IDML)
        info   = get_idml_info(path)
        result = preview_extract(path, lang, max_rows=200)
        # 保留暫存路徑供後續確認匯入
        tmp_id = uuid.uuid4().hex
        tmp_path = os.path.join(UPLOAD_DIR, f'idml_{tmp_id}.idml')
        os.rename(path, tmp_path)
        return jsonify({'ok': True, 'tmp_id': tmp_id, 'info': info, **result})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/import/idml-confirm', methods=['POST'])
def api_import_idml_confirm():
    """確認將 IDML 萃取的文字存入資料庫。"""
    data    = request.json or {}
    tmp_id  = data.get('tmp_id', '')
    lang    = data.get('lang_code', 'ENG')
    product = data.get('product', '')
    chapter = data.get('chapter', '')
    tmp_path = os.path.join(UPLOAD_DIR, f'idml_{tmp_id}.idml')
    if not os.path.exists(tmp_path):
        return jsonify({'ok': False, 'error': '暫存檔案不存在，請重新上傳'}), 400
    try:
        rows = extract_for_import(tmp_path, lang)
        for r in rows:
            r.setdefault('product', product)
            r.setdefault('chapter', chapter)
        stats = db.bulk_upsert(rows)
        os.remove(tmp_path)
        return jsonify({'ok': True, **stats, 'total_rows': len(rows)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ──────────────────────────────────────────
# 匯出
# ──────────────────────────────────────────

@app.route('/api/export/excel')
def api_export_excel():
    """將整個資料庫匯出為 Excel。"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    rows = db.get_all()
    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = '多語言資料庫'

    hfill = PatternFill('solid', fgColor='1B2A4A')
    hfont = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    center = Alignment(horizontal='center', vertical='center')

    all_cols  = ['id', 'product', 'chapter'] + LANG_CODES + ['created_at', 'updated_at']
    ws.append(all_cols)
    for ci, col in enumerate(all_cols, 1):
        cell = ws.cell(1, ci)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = center
        ws.column_dimensions[get_column_letter(ci)].width = (
            5 if col == 'id' else 15 if col in ('product', 'chapter') else
            35 if col in LANG_CODES else 20
        )

    for row in rows:
        ws.append([row.get(c, '') for c in all_cols])

    ws.freeze_panes = 'D2'

    out_path = os.path.join(OUTPUT_DIR, 'translations_export.xlsx')
    wb.save(out_path)
    return send_file(out_path, as_attachment=True,
                     download_name='translations_export.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/export/template')
def api_export_template():
    """下載空白翻譯 Excel 範本。"""
    out = os.path.join(OUTPUT_DIR, 'translation_template.xlsx')
    generate_template(out)
    return send_file(out, as_attachment=True,
                     download_name='translation_template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ──────────────────────────────────────────
# IDML 修正（核心功能）
# ──────────────────────────────────────────

@app.route('/api/patch/upload-idml', methods=['POST'])
def api_patch_upload_idml():
    """上傳待修正的 IDML，回傳暫存 ID。"""
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': '請上傳 IDML'}), 400
    try:
        path   = save_upload(request.files['file'], ALLOWED_IDML)
        tmp_id = uuid.uuid4().hex
        tmp_path = os.path.join(UPLOAD_DIR, f'patch_{tmp_id}.idml')
        os.rename(path, tmp_path)
        info = get_idml_info(tmp_path)
        return jsonify({'ok': True, 'tmp_id': tmp_id, 'info': info,
                        'filename': request.files['file'].filename})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/patch/upload-instructions', methods=['POST'])
def api_patch_upload_instructions():
    """上傳修改指示 Excel/CSV，回傳解析後的指示清單（預覽）。"""
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': '請上傳指示檔案'}), 400
    try:
        path = save_upload(request.files['file'], ALLOWED_INSTR)
        rows = import_excel(path)
        os.remove(path)

        # 解析指示格式：lang_code, find, replace, note
        instructions = []
        for row in rows:
            # 支援標準格式欄位名稱
            lang = row.get('lang_code') or row.get('LANG') or row.get('語言代碼') or row.get('語言') or ''
            find = row.get('find') or row.get('FIND') or row.get('找到原文') or row.get('原文') or ''
            repl = row.get('replace') or row.get('REPLACE') or row.get('修改為') or row.get('修改後') or ''
            note = row.get('note') or row.get('NOTE') or row.get('備註') or ''
            if find and repl:
                instructions.append({'lang_code': lang, 'find': find, 'replace': repl, 'note': note, 'mark_red': True})

        return jsonify({'ok': True, 'instructions': instructions, 'count': len(instructions)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/patch/upload-pdf', methods=['POST'])
def api_patch_upload_pdf():
    """
    接收設計師上傳的確認版 PDF。
    暫存在 outputs 目錄下，檔名為 {run_id}.pdf。
    """
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': '請上傳 PDF 檔案'}), 400
    run_id = request.form.get('run_id', '')
    if not run_id:
        return jsonify({'ok': False, 'error': '未提供 run_id'}), 400

    try:
        file = request.files['file']
        if not file.filename.lower().endswith('.pdf'):
            return jsonify({'ok': False, 'error': '檔案必須是 PDF 格式'}), 400
        
        pdf_path = os.path.join(OUTPUT_DIR, f'{run_id}.pdf')
        file.save(pdf_path)
        return jsonify({'ok': True, 'filename': f'{run_id}.pdf'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/patch/run', methods=['POST'])
def api_patch_run():
    """
    執行 IDML 修正。
    Body: { tmp_id, instructions: [...] }
    回傳: { ok, changes, not_found, idml_file, excel_file }
    """
    data         = request.json or {}
    tmp_id       = data.get('tmp_id', '')
    instructions = data.get('instructions', [])
    orig_name    = data.get('original_filename', 'document.idml')

    idml_path = os.path.join(UPLOAD_DIR, f'patch_{tmp_id}.idml')
    if not os.path.exists(idml_path):
        return jsonify({'ok': False, 'error': '找不到 IDML 檔案，請重新上傳'}), 400

    base_name = os.path.splitext(orig_name)[0]
    run_id    = uuid.uuid4().hex[:8]
    out_idml  = os.path.join(OUTPUT_DIR, f'{base_name}_MARKED_{run_id}.idml')
    out_excel = os.path.join(OUTPUT_DIR, f'{base_name}_report_{run_id}.xlsx')

    try:
        result = patch_idml(idml_path, instructions, out_idml, out_excel)
        layout = extract_layout(out_idml)
        return jsonify({
            'ok': True,
            'changes':    result['changes'],
            'not_found':  result['not_found'],
            'idml_file':  f'{base_name}_MARKED_{run_id}.idml',
            'excel_file': f'{base_name}_report_{run_id}.xlsx',
            'run_id':     run_id,
            'layout':     layout,
        })
    except Exception as e:
        import traceback
        return jsonify({'ok': False, 'error': str(e), 'detail': traceback.format_exc()}), 500


@app.route('/api/patch/verify', methods=['POST'])
def api_patch_verify():
    """
    驗證置換正確性：回讀輸出 IDML，檢查每暉替換的文字是否正確寫入。
    Body: { idml_file, changes: [{find, replace, lang_code, note}, ...] }
    回傳: { ok, results: [{find, replace, lang_code, found, count}] }
    """
    data = request.json or {}
    idml_filename = data.get('idml_file', '')
    changes = data.get('changes', [])

    if not idml_filename:
        return jsonify({'ok': False, 'error': '未提供 IDML 檔案名稱'}), 400

    idml_path = os.path.join(OUTPUT_DIR, idml_filename)
    if not os.path.exists(idml_path):
        return jsonify({'ok': False, 'error': '找不到修改後的 IDML 檔案'}), 400

    import zipfile
    try:
        # 回讀所有 Story，將全部文字合並
        full_text_map = {}  # story_id -> full_text
        with zipfile.ZipFile(idml_path, 'r') as zf:
            from lxml import etree
            for name in zf.namelist():
                if not name.startswith('Stories/'):
                    continue
                story_id = name.replace('Stories/', '').replace('.xml', '')
                raw = zf.read(name)
                try:
                    tree = etree.fromstring(raw)
                    parts = []
                    for el in tree.iter():
                        local = el.tag.split('}')[-1] if '}' in str(el.tag) else el.tag
                        if local == 'Content' and el.text:
                            parts.append(el.text)
                    full_text_map[story_id] = ''.join(parts)
                except Exception:
                    pass

        all_text = '\n'.join(full_text_map.values())

        results = []
        all_pass = True
        for ch in changes:
            replace_text = ch.get('replace', '')
            find_text = ch.get('find', '')
            # 驗證：修改後的文字應在 IDML 中
            count = all_text.count(replace_text) if replace_text else 0
            # 同時驗證原文是否已被清除
            orig_still_present = False
            if find_text != replace_text:
                if find_text in all_text:
                    # 如果原文是新文字的一部分（子字串）
                    if find_text in replace_text:
                        find_count = all_text.count(find_text)
                        # 如果原文出現次數大於新文字出現次數，說明有其他地方殘留了舊原文
                        if find_count > count:
                            orig_still_present = True
                    else:
                        # 原文不是新文字的一部分，但依然存在，說明殘留了
                        orig_still_present = True

            passed = count > 0 and not orig_still_present
            if not passed:
                all_pass = False
            results.append({
                'lang_code': ch.get('lang_code', ''),
                'find': find_text,
                'replace': replace_text,
                'note': ch.get('note', ''),
                'page': ch.get('page', ''),
                'found': count > 0,
                'count': count,
                'orig_still_present': orig_still_present,
                'passed': passed,
            })

        return jsonify({
            'ok': True,
            'all_pass': all_pass,
            'results': results,
            'total': len(results),
            'passed': sum(1 for r in results if r['passed']),
            'failed': sum(1 for r in results if not r['passed']),
        })
    except Exception as e:
        import traceback
        return jsonify({'ok': False, 'error': str(e), 'detail': traceback.format_exc()}), 500


@app.route('/api/patch/export-pdf', methods=['POST'])
def api_patch_export_pdf():
    """
    透過本地 InDesign 將修改後的 IDML 轉存為 PDF 檔案。
    Body: { idml_file }
    """
    data = request.json or {}
    idml_filename = data.get('idml_file', '')
    if not idml_filename:
        return jsonify({'ok': False, 'error': '未提供 IDML 檔案名稱'}), 400

    lang_code = data.get('lang_code', '')

    idml_path = os.path.join(OUTPUT_DIR, idml_filename)
    if not os.path.exists(idml_path):
        return jsonify({'ok': False, 'error': '找不到修改後的 IDML 檔案'}), 400

    base_name = os.path.splitext(idml_filename)[0]
    pdf_filename = f'{base_name}.pdf'
    pdf_path = os.path.join(OUTPUT_DIR, pdf_filename)

    # 執行轉存
    success = export_pdf_via_indesign(idml_path, pdf_path, lang_code=lang_code)
    if success:
        return jsonify({'ok': True, 'pdf_file': pdf_filename})
    else:
        return jsonify({'ok': False, 'error': '呼叫 InDesign 轉存 PDF 失敗，請確認您的 Mac 是否已安裝並啟動 InDesign'}), 500


@app.route('/api/patch/download/<filename>')
def api_patch_download(filename):
    path = os.path.join(OUTPUT_DIR, filename)
    if not os.path.exists(path):
        return jsonify({'error': '檔案不存在'}), 404
    ext = filename.rsplit('.', 1)[-1].lower()
    mime = (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        if ext == 'xlsx' else (
            'application/pdf' if ext == 'pdf' else 'application/octet-stream'
        )
    )
    return send_file(path, as_attachment=True, download_name=filename, mimetype=mime)


@app.route('/api/patch/instruction-template')
def api_instruction_template():
    """下載修改指示 Excel 範本（給 PM 使用）。"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '修改指示'

    hfill  = PatternFill('solid', fgColor='1B2A4A')
    hfont  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    center = Alignment(horizontal='center', vertical='center')
    note_fill = PatternFill('solid', fgColor='FFF8E7')

    headers = ['語言代碼', '找到原文', '修改為', '備註']
    widths  = [12, 55, 55, 25]
    ws.append(headers)
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(1, ci)
        c.fill = hfill
        c.font = hfont
        c.alignment = center
        ws.column_dimensions[get_column_letter(ci)].width = w

    # 示範列
    examples = [
        ('ENG', 'Do not use near water.', 'Do not operate near water.', 'PM-001'),
        ('CHT', '請勿在水邊使用。', '請勿在水邊操作。', 'PM-001'),
        ('GER', 'Nicht in Wassernähe benutzen.', 'Nicht in der Nähe von Wasser betreiben.', 'PM-001'),
    ]
    for row in examples:
        ws.append(list(row))
        ri = ws.max_row
        for ci in range(1, 5):
            ws.cell(ri, ci).fill = note_fill
            ws.cell(ri, ci).font = Font(name='Arial', size=10, italic=True, color='666666')

    # 說明列
    ws.append([])
    ws.append(['＊語言代碼清單：ENG GER DUT DAN FRE SPA ITA GRK POL PRB RUS CHT JPN KOR VTM THI ARB TRK CHS'])
    ws.cell(ws.max_row, 1).font = Font(name='Arial', size=9, color='888888')
    ws.merge_cells(f'A{ws.max_row}:D{ws.max_row}')

    ws.freeze_panes = 'A2'

    out = os.path.join(OUTPUT_DIR, 'PM_修改指示範本.xlsx')
    wb.save(out)
    return send_file(out, as_attachment=True, download_name='PM_修改指示範本.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ──────────────────────────────────────────
# 從資料庫套用語言到 IDML
# ──────────────────────────────────────────

@app.route('/api/apply/run', methods=['POST'])
def api_apply_run():
    """
    從資料庫中以 ENG 原文搜尋，將指定語言套用到 IDML。
    Body: { tmp_id, lang_code, original_filename }
    """
    data      = request.json or {}
    tmp_id    = data.get('tmp_id', '')
    lang      = data.get('lang_code', '')
    orig_name = data.get('original_filename', 'document.idml')
    # 不翻譯清單：該清單內的英文段落將保留英文原文，不套用任何翻譯也不標記顏色
    raw_skip  = data.get('skip_texts', [])  # list of str
    # 正規化：小寫 + strip，方便不區分大小寫比對
    skip_set  = {s.strip().lower() for s in raw_skip if s and s.strip()}

    if not lang:
        return jsonify({'ok': False, 'error': '請選擇目標語言'}), 400

    idml_path = os.path.join(UPLOAD_DIR, f'patch_{tmp_id}.idml')
    if not os.path.exists(idml_path):
        return jsonify({'ok': False, 'error': '找不到 IDML，請重新上傳'}), 400

    # 從 IDML 萃取 ENG 文字，支援軟換行（以 \n 分割後單獨比對），拆分前綴與尾綴後過濾噪點
    stories = extract_stories(idml_path)
    eng_texts = set()
    for story in stories:
        for para in story['paragraphs']:
            # 支援軟換行與特殊換行字元分割
            parts = para['text'].replace('\r', '\n').replace('\u2028', '\n').replace('\u2029', '\n').split('\n')
            for part in parts:
                t = clean_text(part)
                if t:
                    prefix, core, suffix = split_prefix_suffix(t)
                    core_cleaned = clean_text(core)
                    if core_cleaned:
                        if not should_skip_translation(core_cleaned):
                            eng_texts.add(core_cleaned)

    # 查詢資料庫建立替換指示
    instructions = []
    has_any_translation = False
    for eng_text in eng_texts:
        # 不翻譯清單：段落在清單內則直接跳過（不生成任何指令，保留英文原文不標色）
        if skip_set and eng_text.lower() in skip_set:
            continue
        row = db.lookup_eng(eng_text)
        if row and row.get(lang):
            has_any_translation = True
            instructions.append({
                'lang_code':   lang,
                'find':        eng_text,
                'replace':     row[lang],
                'note':        'DB auto-apply',
                'mark_red':    False,   # 黑色：成功套用翻譯
                'mark_green':  False,
                'exact_match': True,    # 只匹配完整段落，防止子字串誤改
            })
        else:
            similar_row = db.find_similar_eng(eng_text)
            if similar_row:
                sim_pct = int(similar_row['similarity'] * 100)
                instructions.append({
                    'lang_code':   lang,
                    'find':        eng_text,
                    'replace':     eng_text,
                    'note':        f'DB Similar Found (ID: {similar_row["id"]}, Similarity: {sim_pct}%): {similar_row["ENG"]}',
                    'mark_red':    False,
                    'mark_green':  False,
                    'mark_orange': True,     # 橘色：疑似錯字/相似句型
                    'exact_match': True,     # 只匹配完整段落，防止子字串誤改
                })
            else:
                instructions.append({
                    'lang_code':   lang,
                    'find':        eng_text,
                    'replace':     eng_text,
                    'note':        'DB Translation Missing (Green)',
                    'mark_red':    False,
                    'mark_green':  True,    # 綠色：翻譯缺失，保留英文
                    'exact_match': True,    # 只匹配完整段落，防止子字串誤改
                })

    if not has_any_translation:
        return jsonify({'ok': False, 'error': f'資料庫中找不到對應的 {lang} 翻譯'}), 400

    base_name = os.path.splitext(orig_name)[0]
    run_id    = uuid.uuid4().hex[:8]
    out_idml  = os.path.join(OUTPUT_DIR, f'{base_name}_{lang}_{run_id}.idml')
    out_excel = os.path.join(OUTPUT_DIR, f'{base_name}_{lang}_report_{run_id}.xlsx')

    try:
        result = patch_idml(idml_path, instructions, out_idml, out_excel)
        layout = extract_layout(out_idml)
        
        # 計算成功替換與翻譯缺失（綠字標記）的統計數字
        applied_count = sum(c.get('count', 1) for c in result['changes'] if not c.get('mark_green', False) and not c.get('mark_red', False) and not c.get('mark_orange', False))
        missing_count = sum(c.get('count', 1) for c in result['changes'] if c.get('mark_green', False))

        return jsonify({
            'ok': True,
            'applied':    applied_count,
            'not_found':  missing_count,
            'idml_file':  os.path.basename(out_idml),
            'excel_file': os.path.basename(out_excel),
            'run_id':     run_id,
            'layout':     layout,
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ──────────────────────────────────────────
# PM 確認流程
# ──────────────────────────────────────────

@app.route('/api/pm-review/generate', methods=['POST'])
def api_pm_review_generate():
    """
    根據修正結果產生 PM 確認用 HTML 檔案。
    Body: { run_id, changes, not_found, original_filename }
    回傳: { ok, html_file }
    """
    data      = request.json or {}
    run_id    = data.get('run_id', '')
    changes   = data.get('changes', [])
    not_found = data.get('not_found', [])
    orig_name = data.get('original_filename', 'document.idml')
    layout    = data.get('layout', None)

    base_name = os.path.splitext(orig_name)[0]
    html_filename = f'PM確認_{base_name}_{run_id}.html'
    out_path  = os.path.join(OUTPUT_DIR, html_filename)

    from datetime import datetime
    try:
        pdf_temp_filename = f'{run_id}.pdf'
        pdf_temp_path = os.path.join(OUTPUT_DIR, pdf_temp_filename)
        has_pdf = os.path.exists(pdf_temp_path)

        pdf_dest_filename = f'{base_name}.pdf'

        meta = {
            'filename':  orig_name,
            'date':      datetime.now().strftime('%Y-%m-%d %H:%M'),
            'run_id':    run_id,
        }
        if has_pdf:
            meta['pdf_filename'] = pdf_dest_filename
            meta['has_pdf'] = True

        generate_pm_review_html(
            changes=changes,
            not_found=not_found,
            output_path=out_path,
            meta=meta,
            layout=layout
        )

        if has_pdf:
            import zipfile
            zip_filename = f'PM確認套件_{base_name}_{run_id}.zip'
            zip_path = os.path.join(OUTPUT_DIR, zip_filename)
            
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                # 寫入 HTML (在 ZIP 中使用友善的檔名)
                html_in_zip = f'PM確認_{base_name}.html'
                zf.write(out_path, arcname=html_in_zip)
                # 寫入 PDF
                zf.write(pdf_temp_path, arcname=pdf_dest_filename)
            
            return jsonify({'ok': True, 'zip_file': zip_filename, 'is_zip': True})
        else:
            return jsonify({'ok': True, 'html_file': html_filename, 'is_zip': False})
    except Exception as e:
        import traceback
        return jsonify({'ok': False, 'error': str(e), 'detail': traceback.format_exc()}), 500


@app.route('/api/pm-review/parse-reply', methods=['POST'])
def api_pm_review_parse_reply():
    """
    解析 PM 回傳的 JSON 確認檔，整理出需要重改的清單。
    Body: 上傳 JSON 檔
    回傳: { ok, summary, needs_redo, confirmed }
    """
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': '請上傳 PM 確認回覆 JSON 檔'}), 400
    try:
        file    = request.files['file']
        content = file.read().decode('utf-8')
        reply   = json.loads(content)

        items    = reply.get('items', [])
        confirmed = [it for it in items if it.get('confirmed') and not it.get('comment', '').strip()]
        with_comments = [it for it in items if it.get('confirmed') and it.get('comment', '').strip()]
        needs_redo   = [it for it in items if not it.get('confirmed')]

        return jsonify({
            'ok':           True,
            'meta':         reply.get('meta', {}),
            'confirmed':    confirmed,
            'with_comments': with_comments,
            'needs_redo':   needs_redo,
            'not_found':    reply.get('not_found', []),
            'summary': {
                'total':         len(items),
                'confirmed':     len(confirmed),
                'with_comments': len(with_comments),
                'needs_redo':    len(needs_redo),
            }
        })
    except json.JSONDecodeError:
        return jsonify({'ok': False, 'error': '無效的 JSON 檔案格式'}), 400
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
# ──────────────────────────────────────────
# 智慧文字提取 API (PDF & Image OCR)
# ──────────────────────────────────────────

HAS_VISION = False
try:
    import objc
    from Cocoa import NSURL
    from Vision import VNImageRequestHandler, VNRecognizeTextRequest
    HAS_VISION = True
except ImportError:
    pass

def detect_language(text):
    import re
    if not text:
        return 'ENG'
    text = text.lower()
    # 1. Quick German accent character check
    if any(c in text for c in ('ä', 'ö', 'ü', 'ß')):
        return 'GER'
        
    # 2. Stop words intersection
    STOP_WORDS = {
        'GER': {'der', 'die', 'das', 'und', 'ist', 'in', 'zu', 'den', 'dem', 'mit', 'von', 'für', 'anzeigefenster', 'steigung'},
        'ENG': {'the', 'and', 'is', 'in', 'to', 'of', 'for', 'on', 'with', 'at', 'incline', 'calories', 'speed', 'time', 'distance'},
        'DUT': {'de', 'het', 'een', 'en', 'van', 'ik', 'te', 'dat', 'die', 'in', 'voor', 'met'},
        'DAN': {'og', 'i', 'jeg', 'det', 'at', 'en', 'den', 'til', 'med', 'på'},
        'FRE': {'le', 'la', 'les', 'et', 'un', 'une', 'en', 'que', 'est', 'dans', 'pour', 'avec'},
        'SPA': {'el', 'la', 'los', 'las', 'un', 'una', 'y', 'es', 'en', 'con', 'para', 'por'},
        'ITA': {'il', 'la', 'i', 'gli', 'le', 'un', 'una', 'e', 'di', 'in', 'per', 'con'},
        'POL': {'w', 'i', 'z', 'na', 'do', 'o', 'że', 'to', 'nie', 'się'},
        'PRB': {'o', 'a', 'os', 'as', 'um', 'uma', 'e', 'em', 'com', 'para', 'por'},
        'RUS': {'и', 'в', 'во', 'не', 'что', 'он', 'на', 'я', 'с', 'со'},
        'TRK': {'bir', 've', 'bu', 'ne', 'da', 'de', 'için', 'ile', 'en', 'o'},
        'GRK': {'και', 'το', 'του', 'τα', 'στην', 'στο', 'την', 'της', 'από', 'που'},
        'VTM': {'và', 'là', 'trong', 'của', 'được', 'một', 'có', 'cho', 'với', 'người'},
        'THI': {'และ', 'ใน', 'เป็น', 'ของ', 'ได้', 'มี', 'การ', 'ที่', 'ให้', 'กับ'},
        'ARB': {'من', 'في', 'على', 'و', 'أن', 'إلى', 'هذا', 'هذه', 'مع', 'كل'},
        'JPN': {'の', 'に', 'は', 'と', 'を', 'た', 'が', 'で', 'て', 'し'},
        'KOR': {'이', '그', '저', '을', '를', '은', '는', '이', '가', '에'},
        'CHT': set("的是在有個這我我們你們他們與或及以於"),
        'CHS': set("的是在有个这我我们你们他们与或及以于")
    }
    
    words = set(re.findall(r'[a-zA-Z\u00C0-\u024F]+', text))
    scores = {}
    for lang, sw in STOP_WORDS.items():
        if lang in ('CHT', 'CHS'):
            chars = set(text)
            scores[lang] = len(chars.intersection(sw))
        else:
            scores[lang] = len(words.intersection(sw))
            
    max_lang = max(scores, key=scores.get)
    if scores[max_lang] > 0:
        return max_lang
    return 'ENG'

def clean_and_validate_extracted_text(text):
    import re
    if not text:
        return None
    text = text.strip()
    
    # 1. Must contain at least one Chinese character, OR at least one word (with letters and length >= 2)
    has_chinese = any('\u4e00' <= c <= '\u9fff' for c in text)
    has_word = re.search(r'\b[a-zA-Z\u00C0-\u024F]{2,}\b', text)
    if not (has_chinese or has_word):
        return None
        
    # 2. Strip leading bullets, layout decorations, and common list items
    cleaned = re.sub(r'^[•▪▲⏰■○●□◇◆▫★☆☞☞▶▷➔➜\-\*\+•\s]+', '', text)
    cleaned = cleaned.strip()
    if not cleaned:
        return None
    return cleaned

@app.route('/api/extract/pdf', methods=['POST'])
def api_extract_pdf():
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': '未提供檔案'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'ok': False, 'error': '檔案名稱為空'}), 400
        
    filename = secure_filename(file.filename)
    tmp_path = os.path.join(UPLOAD_DIR, f"extract_{uuid.uuid4().hex}_{filename}")
    file.save(tmp_path)
    
    try:
        from pypdf import PdfReader
        reader = PdfReader(tmp_path)
        pages_data = []
        for i, page in enumerate(reader.pages):
            text = page.extract_text() or ""
            raw_lines = [line.strip() for line in text.split('\n') if line.strip()]
            paragraphs = []
            current = ""
            for line in raw_lines:
                if not current:
                    current = line
                else:
                    # 智慧拼接條件：若前行無標點句尾，且當前行首非大寫字首，則進行合併
                    ends_with_punctuation = current[-1] in ('.', '?', '!', ':', ';')
                    starts_with_capital = line[0].isupper() if line else False
                    
                    should_merge = not ends_with_punctuation
                    if starts_with_capital:
                        # 檢查是否為首字大寫的完整單字開頭 (例如: User)
                        if len(line) > 1 and line[1].islower():
                            should_merge = False
                            
                    if should_merge:
                        if current.endswith('-'):
                            current = current[:-1] + line
                        else:
                            current += " " + line
                    else:
                        paragraphs.append(current)
                        current = line
            if current:
                paragraphs.append(current)
                
            cleaned_paragraphs = []
            for p in paragraphs:
                cleaned = clean_and_validate_extracted_text(p)
                if cleaned:
                    cleaned_paragraphs.append(cleaned)
                
            page_text = "\n".join(cleaned_paragraphs)
            detected_lang = detect_language(page_text)
            
            pages_data.append({
                'page': i + 1,
                'paragraphs': cleaned_paragraphs,
                'detected_lang': detected_lang
            })
            
        return jsonify({
            'ok': True,
            'pages': pages_data
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

@app.route('/api/extract/ocr', methods=['POST'])
def api_extract_ocr():
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': '未提供檔案'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'ok': False, 'error': '檔案名稱為空'}), 400
        
    if not HAS_VISION:
        return jsonify({'ok': False, 'error': '本機環境不支援 macOS Vision OCR。請在 Mac 上執行專案。'}), 400
        
    filename = secure_filename(file.filename)
    tmp_path = os.path.join(UPLOAD_DIR, f"ocr_{uuid.uuid4().hex}_{filename}")
    file.save(tmp_path)
    
    try:
        url = NSURL.fileURLWithPath_(tmp_path)
        handler = VNImageRequestHandler.alloc().initWithURL_options_(url, None)
        
        results = []
        def completion_handler(req, error):
            if error:
                print("Vision OCR error:", error)
                return
            observations = req.results()
            if observations:
                for obs in observations:
                    candidates = obs.topCandidates_(1)
                    if not candidates:
                        continue
                    text = candidates[0].string()
                    cleaned_text = clean_and_validate_extracted_text(text)
                    if not cleaned_text:
                        continue
                    box = obs.boundingBox()
                    
                    # Convert coordinates (bottom-left to top-left normalized)
                    x = box.origin.x
                    y = 1.0 - (box.origin.y + box.size.height)
                    w = box.size.width
                    h = box.size.height
                    
                    results.append({
                        'text': cleaned_text,
                        'box': [x, y, w, h]
                    })
                    
        ocr_req = VNRecognizeTextRequest.alloc().initWithCompletionHandler_(completion_handler)
        ocr_req.setRecognitionLevel_(0)  # Accurate level
        ocr_req.setRecognitionLanguages_(["zh-Hant", "en-US"])
        
        success, error = handler.performRequests_error_([ocr_req], None)
        if not success:
            return jsonify({'ok': False, 'error': f"Vision OCR failed: {error}"}), 500
            
        combined_text = "\n".join([b['text'] for b in results])
        detected_lang = detect_language(combined_text)
        
        return jsonify({
            'ok': True,
            'detected_lang': detected_lang,
            'blocks': results
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
@app.route('/api/extract/import', methods=['POST'])
def api_extract_import():
    data = request.json or {}
    rows = data.get('rows', [])
    
    imported_count = 0
    conflict_count = 0
    skipped_count = 0
    
    for row in rows:
        eng_val = row.get('ENG', '').strip() if row.get('ENG') is not None else ''
        if not eng_val:
            skipped_count += 1
            continue
            
        product = row.get('product', '') or ''
        chapter = row.get('chapter', '') or ''
        
        row_data = {'product': product, 'chapter': chapter, 'ENG': eng_val}
        for code in LANG_CODES:
            if code != 'ENG':
                row_data[code] = row.get(code, '') or ''
                
        existing = db.lookup_eng(eng_val)
        if not existing:
            db.add(row_data)
            imported_count += 1
        else:
            has_conflict_on_row = False
            conflicting_langs = {}
            non_conflicting_updates = {}
            
            for code in LANG_CODES:
                if code == 'ENG':
                    continue
                import_translation = row_data.get(code, '') or ''
                db_translation = existing.get(code, '') or ''
                
                if import_translation and db_translation:
                    if import_translation != db_translation:
                        has_conflict_on_row = True
                        conflicting_langs[code] = (db_translation, import_translation)
                elif import_translation and not db_translation:
                    non_conflicting_updates[code] = import_translation
                    
            if has_conflict_on_row:
                for lang, (db_val, import_val) in conflicting_langs.items():
                    db.add_pending_conflict(
                        product=product or existing.get('product', ''),
                        chapter=chapter or existing.get('chapter', ''),
                        eng_text=eng_val,
                        lang_code=lang,
                        db_val=db_val,
                        import_val=import_val
                    )
                    conflict_count += 1
            
            if non_conflicting_updates:
                db.update(existing['id'], non_conflicting_updates)
                imported_count += 1
                
    return jsonify({
        'ok': True,
        'imported_count': imported_count,
        'conflict_count': conflict_count,
        'skipped_count': skipped_count,
        'message': f'{imported_count} 筆資料已成功處理；{conflict_count} 個語言翻譯衝突已移至待處理佇列。'
    })


# ──────────────────────────────────────────
# 啟動
# ──────────────────────────────────────────

if __name__ == '__main__':
    print('\n' + '='*50)
    print('  OM 多語言管理系統 啟動中...')
    print('  請在瀏覽器開啟 http://127.0.0.1:5001')
    print('='*50 + '\n')
    app.run(debug=True, port=5001, host='127.0.0.1')


import openpyxl
from openpyxl.cell.rich_text import CellRichText

wb = openpyxl.load_workbook("outputs/test_direct_report.xlsx")
print("Sheets in workbook:", wb.sheetnames)

ws = wb['⚠️ 疑似錯字與相似句型']
print("Warnings Sheet dimensions:", ws.dimensions)

# Iterate warnings rows
for r in range(2, ws.max_row + 1):
    row_vals = [cell.value for cell in ws[r]]
    print(f"\nRow {r}:")
    print(f"  Col 1 (Index): {row_vals[0]}")
    print(f"  Col 2 (Lang):  {row_vals[1]}")
    print(f"  Col 3 (Find):  {row_vals[2]}")
    print(f"  Col 4 (Diff):  {repr(row_vals[3])}") # This should print the CellRichText object representation
    print(f"  Col 5 (DB text): {row_vals[4]}")
    
    # Check details of CellRichText
    diff_cell = ws.cell(r, 4)
    if isinstance(diff_cell.value, CellRichText):
        print("  Col 4 is CellRichText! Text Blocks:")
        for block in diff_cell.value:
            font = block.font
            color_rgb = font.color.rgb if font and font.color else None
            is_bold = font.b if font else False
            print(f"    - Text: {repr(block.text)} | Color: {color_rgb} | Bold: {is_bold}")

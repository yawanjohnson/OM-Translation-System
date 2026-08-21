import openpyxl

rpath = "outputs/BFX26_T3-21 A4 OM_r1_0_A 081326' 測試_report_f7a96244.xlsx"
wb = openpyxl.load_workbook(rpath)
print("Sheet names:", wb.sheetnames)

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"\n--- Sheet: {sname} (Rows: {ws.max_row}) ---")
    for ri in range(1, ws.max_row + 1):
        row_vals = [cell.value for cell in ws[ri]]
        if any(row_vals):
            print(f"  Row {ri}: {row_vals}")

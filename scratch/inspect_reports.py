import os
import openpyxl

output_dir = 'outputs'
reports = [f for f in os.listdir(output_dir) if f.endswith('.xlsx') and 'report' in f]

for r in sorted(reports, key=lambda x: os.path.getmtime(os.path.join(output_dir, x)), reverse=True)[:5]:
    rpath = os.path.join(output_dir, r)
    print(f"\nReport: {r}")
    try:
        wb = openpyxl.load_workbook(rpath)
        print("  Sheets:", wb.sheetnames)
        for sname in wb.sheetnames:
            if '翻譯缺失' in sname or '缺失' in sname or 'Sheet' in sname:
                ws = wb[sname]
                print(f"  --- {sname} (Rows: {ws.max_row}) ---")
                # print first 15 rows of this sheet
                for ri in range(1, min(ws.max_row + 1, 15)):
                    row_vals = [cell.value for cell in ws[ri]]
                    if any(row_vals):
                        print(f"    Row {ri}: {row_vals}")
    except Exception as e:
        print("  Error loading:", e)

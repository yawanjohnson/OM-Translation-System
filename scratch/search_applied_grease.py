import openpyxl

rpath = "outputs/SCH26_180U-21 OM_ENG_042826'_CHT_report_ccd2b02b.xlsx"
wb = openpyxl.load_workbook(rpath, data_only=True)

print("=== SEARCHING EXCEL REPORT FOR 'grease' ===")
for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"\nSheet: {sname}")
    found = 0
    for ri in range(1, ws.max_row + 1):
        row_vals = [cell.value for cell in ws[ri]]
        for val in row_vals:
            if val and isinstance(val, str) and 'grease' in val.lower():
                print(f"  Row {ri}: {row_vals}")
                found += 1
    if found == 0:
        print("  No matches in this sheet")

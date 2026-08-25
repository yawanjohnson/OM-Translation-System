import os
import openpyxl

output_dir = 'outputs'
reports = [f for f in os.listdir(output_dir) if f.startswith('test_apply_CHT_report_') and f.endswith('.xlsx')]

if not reports:
    print("No test reports found.")
else:
    # Get the latest report
    latest_report = max(reports, key=lambda x: os.path.getmtime(os.path.join(output_dir, x)))
    rpath = os.path.join(output_dir, latest_report)
    print(f"Reading latest report: {rpath}")
    wb = openpyxl.load_workbook(rpath, data_only=True)
    
    sheet_name = '❌ 找不到內容' if '❌ 找不到內容' in wb.sheetnames else '❌ IDML找不到'
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n=== NOT FOUND STRINGS (Total Rows: {ws.max_row}) ===")
        for r in range(1, ws.max_row + 1):
            row_vals = [cell.value for cell in ws[r]]
            if any(row_vals):
                print(f"  Row {r}: {row_vals}")
    else:
        print("Sheet '❌ 找不到內容' not found in report.")


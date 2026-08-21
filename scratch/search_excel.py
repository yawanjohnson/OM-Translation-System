import openpyxl

wb = openpyxl.load_workbook("UB_RB ENG WARNING 042726' 合併.xlsx", read_only=True)
print("Sheet names in Excel file:", wb.sheetnames)

keywords = ["pinch", "handlebars", "crank", "reverse", "Schwinn", "Upright"]

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\nScanning sheet: {sheet_name}")
    row_count = 0
    matches = {kw: [] for kw in keywords}
    for row in ws.iter_rows(values_only=True):
        row_count += 1
        for cell in row:
            if cell and isinstance(cell, str):
                for kw in keywords:
                    if kw.lower() in cell.lower():
                        matches[kw].append((row_count, cell))
                        
    print(f"  Total rows scanned: {row_count}")
    for kw, list_matches in matches.items():
        if list_matches:
            print(f"  Keyword '{kw}': {len(list_matches)} matches found (showing first 3)")
            for r_num, text in list_matches[:3]:
                print(f"    Row {r_num}: {repr(text)}")

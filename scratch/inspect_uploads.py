import os
import shutil
import openpyxl
import csv
import json

uploads_dir = "uploads"
print("=== Scanning uploads directory with temp copies ===")

for f in os.listdir(uploads_dir):
    fpath = os.path.join(uploads_dir, f)
    if f == ".gitkeep":
        continue
        
    if f.endswith("_xlsx") or f.endswith(".xlsx"):
        print(f"\nFile: {f} (Excel)")
        temp_path = fpath + "_temp.xlsx"
        try:
            shutil.copy(fpath, temp_path)
            wb = openpyxl.load_workbook(temp_path, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                print(f"  Sheet: {sheet}")
                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    row_str = " ".join([str(x) for x in row if x is not None])
                    if any(k in row_str for k in ["P1", "P2", "P3", "P4", "P5", "MANUAL", "FAT BURN", "CALORIES"]):
                        print(f"    Row {row_idx}: {row}")
            wb.close()
        except Exception as e:
            print(f"  Error reading Excel: {e}")
        finally:
            if os.path.exists(temp_path):
                os.remove(temp_path)
                
    elif f.endswith("_csv") or f.endswith(".csv"):
        print(f"\nFile: {f} (CSV)")
        try:
            with open(fpath, "r", encoding="utf-8", errors="ignore") as file:
                reader = csv.reader(file)
                for row_idx, row in enumerate(reader):
                    row_str = " ".join(row)
                    if any(k in row_str for k in ["P1", "P2", "P3", "P4", "P5", "MANUAL", "FAT BURN", "CALORIES"]):
                        print(f"    Row {row_idx}: {row}")
        except Exception as e:
            print(f"  Error reading CSV: {e}")

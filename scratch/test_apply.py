import json
import os
import shutil
import openpyxl
from io import BytesIO
from app import app

def test_api_apply():
    uploads_dir = 'uploads'
    test_src = 'patch_028b1f39879844ea9aaf442661f73598.idml'
    src_file = os.path.join(uploads_dir, test_src)
    if not os.path.exists(src_file):
        print(f"Test source file {src_file} does not exist.")
        return
        
    test_tmp_id = "test_apply"
    dest_file = os.path.join(uploads_dir, f"patch_{test_tmp_id}.idml")
    shutil.copyfile(src_file, dest_file)
    print(f"Copied test IDML file: {src_file} -> {dest_file}")
    
    # 2. Call the Flask API
    client = app.test_client()
    req_body = {
        'tmp_id': test_tmp_id,
        'lang_code': 'CHT',
        'original_filename': 'test_apply.idml'
    }
    
    response = client.post(
        '/api/apply/run',
        data=json.dumps(req_body),
        content_type='application/json'
    )
    
    print("\nAPI Response Status:", response.status_code)
    res_data = response.json
    print("API Response Body (Summary):")
    if res_data:
        print(f"  ok: {res_data.get('ok')}")
        print(f"  applied: {res_data.get('applied')}")
        print(f"  not_found: {res_data.get('not_found')}")
        print(f"  idml_file: {res_data.get('idml_file')}")
        print(f"  excel_file: {res_data.get('excel_file')}")
    
    # Clean up the test file
    if os.path.exists(dest_file):
        os.remove(dest_file)
        
    if not res_data or not res_data.get('ok'):
        print("API call failed.")
        return
        
    # 3. Read the output Excel report to check if 'grease' was matched
    excel_file = res_data.get('excel_file')
    if excel_file:
        excel_path = os.path.join('outputs', excel_file)
        if os.path.exists(excel_path):
            print(f"\nOpening generated report: {excel_path}")
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            # 3. Read the output Excel report to check if 'grease' or 'paint' or 'heart' was matched
            for sname in wb.sheetnames:
                ws = wb[sname]
                print(f"  Scanning sheet '{sname}' for 'grease' / 'paint' / 'heart'...")
                for r in range(1, ws.max_row + 1):
                    row_vals = [cell.value for cell in ws[r]]
                    for val in row_vals:
                        if val and isinstance(val, str) and any(kw in val.lower() for kw in ('grease', 'paint', 'heart')):
                            print(f"    [{sname}] Row {r}: {row_vals}")
            # Clean up output files (commented out for debugging)
            # os.remove(excel_path)
            # idml_out = os.path.join('outputs', res_data.get('idml_file'))
            # if os.path.exists(idml_out):
            #     os.remove(idml_out)
        else:
            print("Excel report file not found in outputs.")

if __name__ == '__main__':
    test_api_apply()

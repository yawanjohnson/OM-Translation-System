import os
import openpyxl

output_dir = 'outputs'
reports = [f for f in os.listdir(output_dir) if f.startswith('test_apply_CHT_report_') and f.endswith('.xlsx')]

if not reports:
    print("No test reports found.")
else:
    latest_report = max(reports, key=lambda x: os.path.getmtime(os.path.join(output_dir, x)))
    rpath = os.path.join(output_dir, latest_report)
    print(f"Reading report: {rpath}")
    wb = openpyxl.load_workbook(rpath, data_only=True)
    
    for sname in wb.sheetnames:
        ws = wb[sname]
        for r in range(1, ws.max_row + 1):
            row_vals = [cell.value for cell in ws[r]]
            for val in row_vals:
                if val and isinstance(val, str) and 'MOVING' in val:
                    print(f"[{sname}] Row {r}: {row_vals}")
